from pathlib import Path
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from pfp.reporting.portfolio_report import PortfolioReport


EURO = '#,##0.00 [$€-es-ES]'
PERCENT = '0.00%'
SHARES = '0.########'
HEADER_FILL = PatternFill('solid', fgColor='D9EAF7')


class WorkbookWriter:
    def write(self, report: PortfolioReport, path: str | Path) -> Path:
        output = Path(path)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        dashboard = workbook.active
        dashboard.title = "Dashboard"
        self._write_dashboard(dashboard, report)
        self._write_summary(workbook, report)
        self._write_positions(workbook, report)
        self._write_accounts(workbook, report)
        self._write_movements(workbook, report)
        self._write_allocation(workbook, report)
        workbook.save(output)
        return output

    @staticmethod
    def _write_dashboard(sheet, report: PortfolioReport) -> None:
        sheet.append(["PFP — Dashboard"])
        sheet["A1"].font = Font(bold=True, size=18)
        sheet.append([])
        metrics = [
            ("Patrimonio total", report.total_value),
            ("Efectivo", report.cash),
            ("Valor de mercado", report.market_value),
            ("Capital invertido", report.invested),
            ("P/L realizado", report.realized_gain_loss),
            ("P/L no realizado", report.unrealized_gain_loss),
            ("P/L total", report.realized_gain_loss + report.unrealized_gain_loss),
        ]
        for label, value in metrics:
            sheet.append([label, value])
        for row in range(4, 4 + len(metrics)):
            sheet.cell(row, 1).font = Font(bold=True)
            sheet.cell(row, 2).number_format = EURO

        start = 13
        sheet.cell(start, 1, "Clase")
        sheet.cell(start, 2, "Valor")
        allocation = [("RV", report.equity_value), ("RF", report.fixed_income_value), ("Oro", report.gold_value), ("Cripto", report.crypto_value)]
        for index, (label, value) in enumerate(allocation, start=start + 1):
            sheet.cell(index, 1, label)
            sheet.cell(index, 2, value)
            sheet.cell(index, 2).number_format = EURO
        chart = PieChart()
        chart.title = "Distribución por clase"
        chart.add_data(Reference(sheet, min_col=2, min_row=start, max_row=start + len(allocation)), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=start + 1, max_row=start + len(allocation)))
        sheet.add_chart(chart, "D3")
        WorkbookWriter._autosize(sheet)

    @staticmethod
    def _write_summary(workbook: Workbook, report: PortfolioReport) -> None:
        sheet = workbook.create_sheet("Resumen")
        sheet.append(["PFP — Resumen de cartera"])
        sheet["A1"].font = Font(bold=True, size=16)
        sheet.append([])
        sheet.append([])
        rows = [("Patrimonio total", report.total_value), ("Efectivo", report.cash), ("Capital invertido", report.invested), ("Valor de mercado", report.market_value), ("P/L realizado", report.realized_gain_loss), ("P/L no realizado", report.unrealized_gain_loss), ("Renta variable", report.equity_value), ("Renta fija", report.fixed_income_value), ("Oro", report.gold_value), ("Cripto", report.crypto_value)]
        for label, value in rows:
            sheet.append([label, value])
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=2, max_col=2):
            row[0].number_format = EURO
        WorkbookWriter._autosize(sheet)

    @staticmethod
    def _write_positions(workbook: Workbook, report: PortfolioReport) -> None:
        sheet = workbook.create_sheet("Posiciones")
        WorkbookWriter._header(sheet, ["Símbolo", "Nombre", "Clase", "Participaciones", "Invertido", "Precio medio", "Precio mercado", "Valor mercado", "Peso", "P/L"])
        for p in report.positions:
            sheet.append([p.symbol, p.name, p.portfolio_class, p.shares, p.invested, p.average_price, p.market_price, p.market_value, p.weight, p.gain_loss])
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            row[3].number_format = SHARES
            for index in (4, 5, 6, 7, 9): row[index].number_format = EURO
            row[8].number_format = PERCENT
        WorkbookWriter._autosize(sheet)

    @staticmethod
    def _write_accounts(workbook: Workbook, report: PortfolioReport) -> None:
        sheet = workbook.create_sheet("Cuentas")
        WorkbookWriter._header(sheet, ["Cuenta", "Banco/Broker", "Moneda", "Saldo"])
        for a in report.accounts: sheet.append([a.name, a.broker, a.currency, a.balance])
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4): row[0].number_format = EURO
        WorkbookWriter._autosize(sheet)

    @staticmethod
    def _write_movements(workbook: Workbook, report: PortfolioReport) -> None:
        sheet = workbook.create_sheet("Movimientos")
        WorkbookWriter._header(sheet, ["Fecha", "Broker", "Categoría", "Tipo", "Clase", "Símbolo", "Nombre", "Participaciones", "Precio", "Importe", "Comisión", "Impuestos", "Moneda", "Descripción", "Transaction ID"])
        for m in report.movements:
            sheet.append([m.datetime, m.broker, m.category, m.type, m.asset_class, m.symbol, m.name, m.shares, m.price, m.amount, m.fee, m.tax, m.currency, m.description, m.transaction_id])
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            row[0].number_format = 'yyyy-mm-dd hh:mm'
            row[7].number_format = SHARES
            for index in (8, 9, 10, 11): row[index].number_format = EURO
        sheet.freeze_panes = "A2"
        WorkbookWriter._autosize(sheet)

    @staticmethod
    def _write_allocation(workbook: Workbook, report: PortfolioReport) -> None:
        sheet = workbook.create_sheet("Asignación")
        WorkbookWriter._header(sheet, ["Clase", "Objetivo", "Valor actual", "% actual", "Desviación"])
        targets = {"RV": Decimal("0.75"), "RF": Decimal("0.20"), "GOLD": Decimal("0.05"), "CRYPTO": Decimal("0")}
        values = {"RV": report.equity_value, "RF": report.fixed_income_value, "GOLD": report.gold_value, "CRYPTO": report.crypto_value}
        total = report.market_value
        for cls in ("RV", "RF", "GOLD", "CRYPTO"):
            actual = values[cls] / total if total else Decimal("0")
            sheet.append([cls, targets[cls], values[cls], actual, actual - targets[cls]])
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            row[1].number_format = PERCENT
            row[2].number_format = EURO
            row[3].number_format = PERCENT
            row[4].number_format = PERCENT
        chart = BarChart()
        chart.title = "Actual vs objetivo"
        chart.add_data(Reference(sheet, min_col=2, max_col=2, min_row=1, max_row=sheet.max_row), titles_from_data=True)
        chart.add_data(Reference(sheet, min_col=4, max_col=4, min_row=1, max_row=sheet.max_row), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row))
        sheet.add_chart(chart, "G2")
        WorkbookWriter._autosize(sheet)

    @staticmethod
    def _header(sheet, headers) -> None:
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    @staticmethod
    def _autosize(sheet) -> None:
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(width, 45)
