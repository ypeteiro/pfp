from decimal import Decimal

from openpyxl import load_workbook

from pfp.domain.portfolio import Portfolio
from pfp.domain.position import Position
from pfp.excel.workbook import WorkbookWriter
from pfp.reporting.portfolio_report import PortfolioReport


def _report():
    portfolio = Portfolio(cash=Decimal("1000"))
    portfolio.positions["EUNL"] = Position(
        symbol="EUNL",
        name="MSCI World",
        portfolio_class="RV",
        shares=Decimal("2"),
        invested=Decimal("200"),
        market_price=Decimal("120"),
    )
    return PortfolioReport.from_portfolio(portfolio)


def test_workbook_writer_creates_full_workbook(tmp_path):
    output = tmp_path / "portfolio.xlsx"
    WorkbookWriter().write(_report(), output)
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Dashboard", "Resumen", "Posiciones", "Cuentas", "Movimientos", "Asignación"]
    assert workbook["Resumen"]["A4"].value == "Patrimonio total"
    assert workbook["Resumen"]["B4"].value == Decimal("1240")
    assert workbook["Posiciones"]["A2"].value == "EUNL"
    assert workbook["Posiciones"]["D2"].value == Decimal("2")
    assert workbook["Posiciones"]["F2"].value == Decimal("100")
    assert workbook["Posiciones"]["H2"].value == Decimal("240")
    assert workbook["Asignación"]["A2"].value == "RV"
    assert workbook["Asignación"]["B2"].value == Decimal("0.75")


def test_workbook_writer_creates_parent_directory(tmp_path):
    output = tmp_path / "exports" / "portfolio.xlsx"
    result = WorkbookWriter().write(_report(), output)
    assert result == output
    assert output.exists()
